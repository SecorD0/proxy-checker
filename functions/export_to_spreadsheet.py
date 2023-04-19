import logging
import platform
from typing import Optional

from colorama import Fore, Style
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from data.config import RESULTS_FILE
from utils.db_api.database import db
from utils.db_api.models import Proxy, IPInfo, ProviderInfo, SiteAccessibility
from utils.miscellaneous.cprint import cprint


async def fill_out_sheet(spreadsheet: Workbook, sheet_name: str, db_table: str, rows: list,
                         headers: Optional[list] = None) -> Workbook:
    if sheet_name not in spreadsheet:
        spreadsheet.create_sheet(sheet_name)

    if not headers:
        headers = enumerate(['n'] + list(db.execute(f'SELECT * FROM {db_table}').keys())[1:])

    else:
        headers = enumerate(headers)

    sheet: Worksheet = spreadsheet[sheet_name]
    for column, header in headers:
        cell = sheet.cell(row=1, column=column + 1)
        cell.value = header
        cell.alignment = Alignment(horizontal='center')

    for n_row, row in enumerate(rows):
        sheet.cell(row=n_row + 2, column=1).value = n_row + 1
        for column, value in enumerate(row[1:]):
            sheet.cell(row=n_row + 2, column=column + 2).value = value

    return spreadsheet


async def export_to_spreadsheet() -> None:
    db_table = 'proxies'
    proxies = list(db.execute(f'SELECT * FROM {db_table}'))
    if proxies:
        try:
            spreadsheet = await fill_out_sheet(
                spreadsheet=Workbook(), sheet_name='Proxies', db_table=db_table, rows=proxies
            )
            db_table = 'ips_info'
            ips_info = list(db.execute(f'SELECT * FROM {db_table}'))
            if ips_info:
                spreadsheet = await fill_out_sheet(
                    spreadsheet=spreadsheet, sheet_name='IPs info', db_table=db_table, rows=ips_info
                )

            db_table = 'providers_info'
            providers_info = list(db.execute(f'SELECT * FROM {db_table}'))
            if providers_info:
                spreadsheet = await fill_out_sheet(
                    spreadsheet=spreadsheet, sheet_name='Providers info', db_table=db_table, rows=providers_info
                )

            db_table = 'sites_accessibility'
            db_sites_accessibility = list(db.execute(f'SELECT * FROM {db_table}'))
            if db_sites_accessibility:
                sites_accessibility = {}
                for row in db_sites_accessibility:
                    ip = row[1]
                    if ip in sites_accessibility:
                        sites_accessibility[ip].append(row)

                    else:
                        sites_accessibility[ip] = [row]

                headers = [row[0] for row in list(db.execute(f'SELECT site_url FROM {db_table} GROUP BY site_url'))]
                headers = ['n', 'ip'] + headers
                value_dicts = []
                for key, value in sites_accessibility.items():
                    row = {'ip': key}
                    for data in value:
                        row[data[2]] = data[3]

                    value_dicts.append(row)

                rows = []
                for value_dict in value_dicts:
                    row = []
                    for header in headers:
                        if header in value_dict:
                            row.append(value_dict[header])

                        else:
                            row.append(0)

                    rows.append(row)

                spreadsheet = await fill_out_sheet(
                    spreadsheet=spreadsheet, sheet_name='Sites accessibility', db_table=db_table, rows=rows,
                    headers=headers
                )

            del spreadsheet['Sheet']
            spreadsheet.save(RESULTS_FILE)
            for instance in db.all(Proxy) + db.all(IPInfo) + db.all(ProviderInfo) + db.all(SiteAccessibility):
                db.s.delete(instance)

            db.commit()
            text = '\nResults exported to the result.xlsx file!' if platform.system() == 'Windows' else \
                f'\nResults exported to the {Fore.LIGHTGREEN_EX}result.xlsx{Style.RESET_ALL} file!'
            print(text)

        except BaseException as e:
            logging.exception('export_to_spreadsheet')
            await cprint(text=f'Failed to export results: {e}', color=Fore.RED)

    else:
        await cprint(text='There are no proxies in the DB!', color=Fore.RED)
